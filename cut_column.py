import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter import Tk, filedialog, messagebox
import os

# List of columns to keep (case-sensitive match to Excel headers)
columns_to_keep = [
    "pam_code", "Client_No", "Port", "product", "status", "mode", "responsible", "OA",
    "vat_principal", "int_principal", "DOC_NO", "Doc_Type", "Pay_Date", "EFF_Date",
    " Cash_inflow ", " payment_part_novat ", "name", "type", "pam_code.1",
    "old_contract_number", "portfolio", " vat ", " cr_cost8 ", " dr_int ", " cr_int ",
    " gain ", " pt_40 ", " cr_int+gain ", " check ", "year_eff", "month_eff",
    "vat_from_cash_(non-loan)", "vat_from_direct_sell", "vat_from_auction_sell"
]

def select_file():
    root = Tk()
    root.withdraw()
    return filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )

def Cut_column_summary(source_dir):
    file_path = source_dir
    if not file_path:
        print("No file selected.")
        return

    try:
        df = pd.read_excel(file_path, sheet_name="transaction_record", engine="openpyxl")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open the file or sheet:\n{e}")
        return

    # Save a modified version with all data first (preserve formatting)
    output_file = os.path.splitext(file_path)[0] + "-cut.xlsx"
    df.to_excel(output_file, sheet_name="transaction_record", index=False)

    # Open with openpyxl to hide unwanted columns
    wb = load_workbook(output_file)
    ws = wb["transaction_record"]

    # Get actual headers from Excel
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    for idx, header in enumerate(headers):
        if header not in columns_to_keep:
            col_letter = get_column_letter(idx + 1)
            ws.column_dimensions[col_letter].hidden = True

    wb.save(output_file)
    #messagebox.showinfo("Done", f"Saved file with hidden columns:\n{output_file}")