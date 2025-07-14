import pandas as pd
from tkinter import Tk, filedialog, messagebox, Toplevel, Text, Scrollbar, RIGHT, Y, BOTH, END
import msoffcrypto
import openpyxl
from io import BytesIO
import win32com.client as win32
import os

def select_file(title):
    return filedialog.askopenfilename(title=title, filetypes=[("Excel Files", "*.xlsx *.xls")])

def main():
    root = Tk()
    root.withdraw()

    # Select files
    file1_path = select_file("Select the Data Daily File (to be updated)")
    if not file1_path:
        messagebox.showerror("Error", "No daily data file selected.")
        return

    file2_path = select_file("Select the Update Data File")
    if not file2_path:
        messagebox.showerror("Error", "No update data file selected.")
        return
        
    # Decrypt file
    decrypted = BytesIO()
    with open(file1_path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password="pam2025")  # Use your password
        office_file.decrypt(decrypted)
        
    try:
        # Load both files
        df_update = pd.read_excel(file2_path, sheet_name="ALL")
        update_dict = df_update.set_index("เลขที่สัญญาใหม่").to_dict("index")
        
        # Load workbook with formulas (original file with password removed)
        wb = openpyxl.load_workbook(decrypted)
        ws = wb["Daily_Report"]
        
        # Define mappings between column headers
        mapping = {
            "product": "สินเชื่อ",
            "QMC_status": "Current Status",
            "responsibility": "Owner",
            "oa": "OA",
            "assign_status": "Assign Status",
            "note": "Assign Note",
            "assign_date": "Assign Date"
        }
        
        # Get column indexes for mapping (1-based)
        headers = [cell.value for cell in ws[1]]
        col_indexes = {name: headers.index(name) + 1 for name in mapping}
        
        # Get index of ID column in Daily_Report
        id_col_index = headers.index("pam_code") + 1
        
        # Loop through rows to update
        for row in ws.iter_rows(min_row=2):  # Skip header
            pamcode = row[id_col_index - 1].value
            if pamcode in update_dict:
                update_values = update_dict[pamcode]
                for col_name, update_col_name in mapping.items():
                    col_idx = col_indexes[col_name]
                    new_value = update_values.get(update_col_name, "")
                    row[col_idx - 1].value = new_value  # Preserve formulas in other columns
        
        # Preview (e.g., print first 5 rows to confirm)
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6), start=2):
            print(f"Row {i}: {[cell.value for cell in row]}")
        
        # Save updated file (e.g., no password first)
        wb.save("updated_temp.xlsx")
        
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(os.path.abspath("updated_temp.xlsx"))
        wb.SaveAs(os.path.abspath(file1_path), Password="pam2025")
        wb.Close()
        
        # Step 6: Clean up
        os.remove("updated_temp.xlsx")
        print(f"Updated and saved to: {file1_path}")

    except Exception as e:
        messagebox.showerror("Error", str(e))

if __name__ == "__main__":
    main()